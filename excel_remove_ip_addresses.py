#/usr/bin/python3

#current issues:
#if sheet has duplicates of multiple IPs, it only removes one
#range has to be set manually; could a while loop be used instead
#need to break everything into functions

import openpyxl #module for working with excel files

wb = openpyxl.load_workbook('/mnt/c/temp/test.xlsx') #load the workbook with IPs to clean up
sheet = wb['Sheet1'] #set the active sheet; only works with default single sheet for now
ipremove_list = input("Enter exact path of list of IPs to remove: ") #location of list of IPs that should be removed from sheet

try:
    with open(ipremove_list) as ipremovelist_file:
        for line in ipremovelist_file:
            for i in range(1, 20000):
                line = line.strip() #built in method to remove whitespace from strings
                if sheet.cell(row=i, column=1).value == line: #find each instance of IPs to remove starting at row i (1) and going through range
                    print("Found IP", i)
                    sheet.delete_rows(i) #delete the rows with the found IP
except: #generic error catching
    print("Something does not work")

wb.save('/mnt/c/temp/testupdate3.xlsx') #save result as new excel file


#for i in range(1, 8):
    #print(i, sheet.cell(row=i, column=1).value)
#    if sheet.cell(row=i, column=1).value == '10.10.0.3':
        #print("found")
        #sheet.cell(row=i, column=1).value = None
        #print("New value:", sheet.cell(row=i, column=1).value)
        sheet.delete_rows(i)
#    else:
        #print("Nothing found")
#        continue
    #    sheet.cell(row=i, column=1).value = None
    #    print("Found value", sheet.cell(row=i, column=1).value
