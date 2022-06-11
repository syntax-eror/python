#/usr/bin/python3

import openpyxl #module for working with excel files

def read_workbook():
    try:
        input_workbook = input("Enter exact path of Excel file to read: ")
        wb = openpyxl.load_workbook(input_workbook)
    except:
        print("Unable to open provided file, check path and filename and try again")
        input("Press <ENTER> to exit")
    return(wb)
	
def remove_ips(wb):
    sheet = wb.active #select last active sheet in workbook, if more than one sheet exists
    try:
        ipremove_list = input("Enter exact path of IP remove list: ")
    except:
        print("Unable to open provided file, check path and filename and try again")
        input("Press <ENTER> to exit")
    try:
        with open(ipremove_list) as ipremovelist_file:
            for line in ipremovelist_file:
                for i in range(1, sheet.max_row + 1): #adjust for missing last line
                    line = line.strip() #built in method to remove whitespace from strings
                    if sheet.cell(row=i, column=1).value == line: #find each instance of IPs to remove starting at row i (1) and going through range
                        sheet.delete_rows(i) #delete the rows with the found IP
    except: #generic error catching
        print("Something does not work")
		
def save_workbook(wb):
    save_location = input("Enter exact path of file to save to: ")
    wb.save(save_location)
    print("Output saved to" save_location)
    input("Press <ENTER> to exit")

wb = read_workbook()
remove_ips(wb)
save_workbook(wb)
