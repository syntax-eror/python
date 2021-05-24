#!/usr/bin/env python3

import openpyxl #pip3 install openpyxl

try:
    location = input("Enter location of file:")
except:
    print("Invalid location double check")

wb = openpyxl.load_workbook(location) #replace with var
sheet = wb['Sheet1'] #can call specific sheets in a wb; Sheet1 is the default
cella1 = sheet['A1']
print(cella1.value)
sheet['A1'] = 1 #assign value directly to cell
sheet['A2'] = None #set value to None

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=1).value)
    
for i in range(1, 8):
    if sheet.cell(row=i, column=1).value == '172.16.39.1':
        sheet.cell(row=i, column=1).value = None
    else:
        print(i, sheet.cell(row=i, column=1).value)
