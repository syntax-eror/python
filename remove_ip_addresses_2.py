#/usr/bin/python3

#current issues:
#if sheet has duplicates of multiple IPs, it only removes one
#range has to be set manually; could a while loop be used instead?

#to do:
#remove duplicates?
#set up way to read from xlsx file instead of txt file
#need to break everything into functions

import openpyxl #module for working with excel files

def remove_ips(wb, sheet, ipremove_list):
    #try:
    with open(ipremove_list) as ipremovelist_file:
        for line in ipremovelist_file:
            for i in range(1, sheet.max_row):
                line = line.strip() #built in method to remove whitespace from strings
                if sheet.cell(row=i, column=1).value == line: #find each instance of IPs to remove starting at row i (1) and going through range
                    print("Found IP", i)
                    sheet.delete_rows(i) #delete the rows with the found IP
    #except: #generic error catching
        #print("Something does not work")
    wb.save('/mnt/c/temp/test/testupdate1.xlsx') #save result as new excel file


wb = openpyxl.load_workbook('/mnt/c/temp/test/test.xlsx') #load the Excel file with IPs to clean up
sheet = wb['Sheet1'] #set the active sheet; only works with default single sheet Sheet1 for now
ipremove_list = '/mnt/c/temp/test/ip.txt'

remove_ips(wb, sheet, ipremove_list)
